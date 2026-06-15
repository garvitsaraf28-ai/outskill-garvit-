#!/usr/bin/env python3
"""OutSkill new-joiner -> sales-ready user journey (v2, merged & improved).

Combines the code-mapped journey with the welcome dashboard, shadowing/
live-support transition, sharper feedback model, sitemap and MVP.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------- palette
NAVY   = "1F3864"
BLUE   = "2E75B6"
fills = {
    "g": PatternFill("solid", fgColor="F2F2F2"),   # access / welcome
    "a": PatternFill("solid", fgColor="E2EFDA"),   # company onboarding (green)
    "b": PatternFill("solid", fgColor="FCE4D6"),   # sales knowledge + practice (orange)
    "y": PatternFill("solid", fgColor="FFF2CC"),   # readiness / certification (yellow)
    "l": PatternFill("solid", fgColor="DDEBF7"),   # live / real work (blue)
}
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FILL  = PatternFill("solid", fgColor=BLUE)

TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
SUB_FONT   = Font(italic=True, color="FFFFFF", size=10)
HEAD_FONT  = Font(bold=True, color="FFFFFF", size=11)
SEC_FONT   = Font(bold=True, color=NAVY, size=12)
CELL_FONT  = Font(size=10)
BOLD_FONT  = Font(bold=True, size=10)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CTR  = Alignment(wrap_text=True, vertical="center", horizontal="center")

wb = Workbook()


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font, c.fill, c.alignment = TITLE_FONT, TITLE_FILL, CTR
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font, c.fill, c.alignment = SUB_FONT, TITLE_FILL, CTR
    ws.row_dimensions[2].height = 18


def subheader(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.alignment = SEC_FONT, Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[row].height = 24
    return row + 1


def minihead(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = HEAD_FONT, HEAD_FILL, CTR, BORDER
    ws.row_dimensions[row].height = 24
    return row + 1


def minirows(ws, row, rows, tone="g", bold0=True, heights=None):
    for r, rw in enumerate(rows):
        rr = row + r
        for c, val in enumerate(rw):
            cell = ws.cell(row=rr, column=c + 1, value=val)
            cell.fill, cell.border, cell.alignment = fills[tone], BORDER, WRAP
            cell.font = BOLD_FONT if (bold0 and c == 0) else CELL_FONT
        ws.row_dimensions[rr].height = (heights[r] if heights else 30)
    return row + len(rows) + 1   # +1 blank spacer


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ================================================================ SHEET 1: SNAPSHOT
s1 = wb.active
s1.title = "1. Journey Snapshot"
title_block(s1, "OutSkill Training Journey — In Short (v2)",
            "Day 1 new joiner  ->  Sales Ready  ->  on-ground live work  ->  continuous improvement", 5)
r = minihead(s1, 4, ["Stage", "Stage Name", "When", "In Short", "App Section"])
snap = [
    ("Stage 0", "Sign up / Login", "Day 1",
     "Logs in with company email; role + track (generalist/engineering) detected; profile + personalised dashboard created.",
     "cover + auth", "g"),
    ("Stage 1", "Welcome Dashboard", "Day 1",
     "NOT training yet: 'Welcome to OutSkill', progress bar, today's task, roadmap, 'Start Day 1' button.",
     "dashboard (new)", "g"),
    ("Stage 2", "Company Foundations", "Day 1",
     "What OutSkill is, the products, the TOFU/MOFU/BOFU funnel, the two masterminds + a short quiz.",
     "foundations (new)", "a"),
    ("Stage 3", "Mastermind Immersion", "Days 2-5",
     "Watch the recent Generalist + Engineering recordings to learn HOW we pitch the Accelerator.",
     "mastermind (new)", "a"),
    ("Stage 4", "Department Overview", "Day 5-6",
     "See Sales / Marketing / Operations, then pick the Inside Sales track.",
     "dept", "a"),
    ("Stage 5", "Product & Offer Training", "Week 1",
     "Master the Accelerator: value, pricing & EMI, who it's for, and why it beats bootcamp.",
     "home / salesuccess", "b"),
    ("Stage 6", "Sales Process Training", "Week 1-2",
     "Learn the 8-step call flow: open -> discover -> diagnose -> objections -> pitch -> close -> follow-up.",
     "home / salesuccess", "b"),
    ("Stage 7", "Mock-Call Room", "Week 2+",
     "Practice vs AI personas; 4 hint-modes (Beginner->Senior) x 6 difficulty levels (Peer->Director).",
     "practice", "b"),
    ("Stage 8", "Feedback Every Call", "Week 2+",
     "Instant feedback: what you said vs what you SHOULD have said, score /100, the next drill.",
     "feedback / reports", "b"),
    ("Stage 9", "Certification / Readiness", "When ready",
     "Pass company quiz + product quiz + pitch test + mock-call score -> become 'Sales Ready'.",
     "salesuccess", "y"),
    ("Stage 10", "Shadowing & Live Support", "Transition",
     "Safe ramp to live: listen -> observe -> co-pilot -> supported live -> independent calls.",
     "(new)", "l"),
    ("Stage 11", "Live Work Mode", "On ground",
     "Assigned leads, follow-ups, scripts, call history, conversion rate, coach feedback.",
     "realcall / followups / reports", "l"),
    ("Stage 12", "Improvement Loop", "Ongoing",
     "Call -> transcript analysed -> AI feedback -> drill one weak area -> repeat mock -> manager sees progress.",
     "(new) + reports", "l"),
]
for row in snap:
    r = minirows(s1, r, [list(row[:-1])], tone=row[-1], bold0=False,
                 heights=[42]) - 1   # no spacer between snapshot rows
    s1.cell(row=r - 1, column=1).font = BOLD_FONT
    s1.cell(row=r - 1, column=2).font = BOLD_FONT
widths(s1, [10, 24, 14, 62, 32])
s1.freeze_panes = "A5"


# ================================================================ SHEET 2: STEP BY STEP
s2 = wb.create_sheet("2. Step-by-Step (Detailed)")
title_block(s2, "OutSkill Training Journey — Step by Step (Full Explanation, v2)",
            "Every step from Day 1 to on-ground work + the continuous-improvement loop", 7)
r = minihead(s2, 4, ["Stage", "Step", "When", "What the Joiner Does",
                     "Detailed Explanation", "Unlocks Next When", "App Section"])
steps = [
    ("Stage 0\nAccess", "0.1 Login w/ company email", "Day 1",
     "Signs up / logs in with @outskill.com email.",
     "Verifies the company email, detects role (new joiner / sales learner / manager / trainer) and captures name, department, cohort/date of joining, current level, and generalist vs engineering track. Roles matter later: seniors/managers can play the customer in higher rounds and view scorecards.",
     "Email verified.", "auth (new)", "g"),
    ("Stage 1\nWelcome", "1.1 Welcome dashboard", "Day 1",
     "Lands on a welcome screen, not training.",
     "First screen answers 3 things instantly — Where am I? ('You're in OutSkill onboarding'), What should I do today? ('Complete Company Basics - Day 1'), How do I become sales-ready? Shows a progress bar, today's task, the training roadmap and a 'Start Day 1' button so the joiner never feels lost.",
     "Clicks 'Start Day 1'.", "dashboard (new)", "g"),
    ("Stage 2\nCompany Foundations", "2.1 What OutSkill is + products", "Day 1",
     "Reads/watches the company overview.",
     "OutSkill is an edtech company that SELLS AI courses. Products: Bootcamp (low ticket), ACCELERATOR (14-day, high ticket, the MAIN offer), Fellowship & Catalyst (6-month).",
     "—", "foundations (new)", "a"),
    ("Stage 2\nCompany Foundations", "2.2 The funnel + masterminds", "Day 1",
     "Learns the funnel structure and the two masterminds.",
     "TOFU = Mastermind + Bootcamp; MOFU = Accelerator; BOFU = Fellowship + Catalyst. Audience: campaigns -> TOFU -> Accelerator -> BOFU. Catalyst needs Accelerator OR Fellowship first. Masterminds: Generalist (Sat/Sun, non-tech) & Engineering (Fri/Sat, Python/tech), ~12-14 hrs, free, where we pitch Bootcamp + Accelerator.",
     "Funnel quiz passed -> Stage 3.", "foundations (new)", "a"),
    ("Stage 3\nMastermind Immersion", "3.1 Watch the recordings", "Days 2-5",
     "Watches recent Generalist (Days 2-3) then Engineering (Days 4-5) recordings, chunked into modules.",
     "Each chunk has key-takeaway notes and 'what was pitched / how it was framed' annotations, so the joiner learns the REAL pitch and can tell which prospect routes to which track.",
     "Both tracks + checks done -> Stage 4.", "mastermind (new)", "a"),
    ("Stage 4\nDepartment Overview", "4.1 Departments + pick track", "Day 5-6",
     "Sees Sales / Marketing / Operations intros, picks Inside Sales.",
     "Replaces the one-by-one department presentations. For Sales it shows: what the team does, who the audience is (mastermind attendees who did NOT pay), what is sold, the rep's role, and how leads come in.",
     "Department chosen -> Stage 5.", "dept (built)", "a"),
    ("Stage 5\nProduct & Offer", "5.1 Accelerator deep dive", "Week 1",
     "Studies the actual offer until they can explain it simply.",
     "Accelerator overview, why it's the main high-ticket offer, pricing & value, who it's for, what differentiates it from bootcamp, and the common objections. Pricing/EMI ties to the payment-gateways sheet (Razorpay / Pine Labs / Shopse / Fibe-Propel / XP). COMPLIANCE: never promise a job or specific salary.",
     "—", "home/salesuccess (new)", "b"),
    ("Stage 6\nSales Process", "6.1 The call flow", "Week 1-2",
     "Learns the end-to-end sales flow.",
     "8-step flow: (1) warm intro (2) why they attended the mastermind (3) their goals (4) why they didn't pay (5) address objections (6) pitch the Accelerator (7) ask for commitment (8) follow-up. Plus lead qualification, discovery questions and handling no-payment users.",
     "Sales knowledge check passed -> Stage 7.", "home/salesuccess (new)", "b"),
    ("Stage 7\nMock-Call Room", "7.1 Practice modes + levels", "Week 2+",
     "Does mock calls vs the AI customer.",
     "4 hint-modes (Beginner script-assisted -> Standard -> Advanced -> Senior no-hints) across 6 difficulty levels (Peer -> Senior -> Utmost Senior -> Manager -> Management -> Director). A real senior/manager can replace the AI in higher rounds. See sheet '3. Mock-Call Room'.",
     "Pass each level's score to climb.", "practice (built)", "b"),
    ("Stage 8\nFeedback", "8.1 Feedback after every call", "Week 2+",
     "Reviews instant feedback and the next drill.",
     "Not just 'good/bad': it shows what went well, what went wrong, missed opportunities, exact better lines ('you said X -> say Y'), a score /100 by criteria, and the single skill+persona to practise next. See sheet '4. Feedback, Readiness & Live'.",
     "—", "feedback/reports (built)", "b"),
    ("Stage 9\nCertification", "9.1 Readiness check", "When ready",
     "Takes the readiness test to unlock live calls.",
     "Must pass: company knowledge quiz + product knowledge quiz + sales pitch test + a minimum mock-call score (incl. the Director round). Passing = 'Sales Ready' badge, logged for managers.",
     "All gates passed -> Stage 10.", "salesuccess (built)", "y"),
    ("Stage 10\nShadowing", "10.1 Safe ramp to live", "Transition",
     "Goes through a supported transition before solo calls.",
     "Listen to recorded senior calls -> observe live customer handling -> co-pilot calls (senior leads) -> supported live calls (rep leads, senior on standby) -> independent calls. Makes onboarding safer and more realistic.",
     "Cleared by senior/manager -> Stage 11.", "(new)", "l"),
    ("Stage 11\nLive Work", "11.1 On-ground work dashboard", "On ground",
     "Handles real prospects day to day.",
     "Sees assigned leads, pending follow-ups, call scripts, call history, performance score, conversion rate and coach feedback. Makes calls, logs outcomes, updates customer status, reviews recordings.",
     "— (ongoing)", "realcall/followups/reports (built)", "l"),
    ("Stage 12\nImprovement Loop", "12.1 Continuous improvement", "Ongoing",
     "Closes the loop on every real call.",
     "Call happens -> transcript analysed -> AI gives feedback -> rep drills one weak area in the Mock-Call Room -> repeats -> score improves -> manager sees progress. This is what keeps the platform useful long-term.",
     "— (continuous)", "(new) + reports", "l"),
]
start = r
for i, row in enumerate(steps):
    rr = start + i
    for c, val in enumerate(row[:-1]):
        cell = s2.cell(row=rr, column=c + 1, value=val)
        cell.fill, cell.border, cell.alignment = fills[row[-1]], BORDER, WRAP
        cell.font = BOLD_FONT if c in (0, 1) else CELL_FONT
    s2.row_dimensions[rr].height = 96
widths(s2, [20, 28, 12, 34, 66, 30, 28])
s2.freeze_panes = "A5"


# ================================================================ SHEET 3: MOCK-CALL ROOM
s3 = wb.create_sheet("3. Mock-Call Room")
title_block(s3, "Stage 7 — Mock-Call Room",
            "Built on the existing practice engine (10 AI personas + 0-100 scorecard)", 4)
r = subheader(s3, 4, "A. Practice modes (how much help)", 4)
r = minihead(s3, r, ["Mode", "Hints", "Purpose"])
r = minirows(s3, r, [
    ["Beginner", "Script-assisted (on-screen prompts)", "First calls — build confidence with a safety net"],
    ["Standard", "Free-flow, no script", "Normal practice conversation"],
    ["Advanced", "Difficult persona, minimal help", "Stretch discovery & objection handling"],
    ["Senior", "No hints, real pressure", "Simulates the senior / manager / director round"],
], tone="a", heights=[26, 26, 26, 26])

r = subheader(s3, r, "B. Difficulty levels (who you face)", 4)
r = minihead(s3, r, ["Level", "Real-World Round", "AI Difficulty", "Example Personas"])
r = minirows(s3, r, [
    ["Level 1", "Peer practice", "Warm / easy", "Priya, Karthik"],
    ["Level 2", "Senior round", "Analytical / skeptical", "Suresh, Meera"],
    ["Level 3", "Utmost senior", "Overconfident / noncommittal", "Aditya, Rahul"],
    ["Level 4", "Manager round", "Hard, hidden blockers", "Vikram, Arjun"],
    ["Level 5", "Management round", "Mixed, high pressure", "Custom + escalated"],
    ["Level 6", "Director round", "Hardest objections", "Custom + escalated"],
], tone="b", heights=[24] * 6)

r = subheader(s3, r, "C. Customer personas & their real blocker", 4)
r = minihead(s3, r, ["Persona / Type", "Stated Objection", "Real Blocker (reveal if probed)"])
r = minirows(s3, r, [
    ["Interested, not ready (Priya)", "\"It feels expensive right now.\"", "Needs permission to invest in herself; can afford EMI."],
    ["No money now (Karthik)", "\"I genuinely can't afford it.\"", "Real budget constraint — needs EMI or a later batch."],
    ["No urgency (Rahul)", "\"Let me think — just send details.\"", "No reason to decide this week; not a real 'no'."],
    ["Came for knowledge only", "\"I attended just to learn, not buy.\"", "Never framed it as a buying decision."],
    ["Offer not relevant (Vikram)", "\"I run a company, no 90 hrs.\"", "Unsure of concrete ROI for HIS business."],
    ["Needs family approval", "\"I have to discuss with family.\"", "Decision isn't fully theirs / quiet money fear."],
    ["Wants to compare (Meera)", "\"How is this different from X?\"", "Needs proof of differentiation & credibility."],
    ["Skeptical of value (Suresh)", "\"Is the certificate worth it?\"", "Needs credibility proof before spending."],
    ["Self-learner (Aditya/Arjun)", "\"I can self-learn / already use ChatGPT.\"", "Doesn't see the gap to building real AI systems."],
], tone="g", heights=[30] * 9)
widths(s3, [30, 36, 46, 20])
s3.freeze_panes = "A4"


# ================================================================ SHEET 4: FEEDBACK / READINESS / LIVE
s4 = wb.create_sheet("4. Feedback, Readiness & Live")
title_block(s4, "Feedback, Readiness, Shadowing & the Improvement Loop", "Stages 8 - 12", 3)
r = subheader(s4, 4, "A. What the AI scores (rubric, 0-100)", 3)
r = minihead(s4, r, ["Criteria", "What 'good' looks like"])
r = minirows(s4, r, [
    ["Opening & rapport", "Warm, clear intro; sets the agenda for the call."],
    ["Discovery & listening", "Asks why they attended and their goals; listens more than talks."],
    ["Pitch quality", "Explains the Accelerator simply, tailored to their goal."],
    ["Objection handling", "Empathy + one specific fact, confirms it landed, finds the real blocker."],
    ["Close & next step", "Clear ask, trial close, secures a concrete next step."],
    ["Confidence & tone", "Calm, confident, empathetic; never pushy."],
    ["Compliance", "Never promises a job or a specific salary."],
], tone="b", heights=[26] * 7)

r = subheader(s4, r, "B. Feedback after every call", 3)
r = minihead(s4, r, ["Element", "Detail"])
r = minirows(s4, r, [
    ["What went well", "Specific strengths from this call."],
    ["What went wrong", "Specific misses, with the moment they happened."],
    ["Missed opportunities", "Probes or closes the rep skipped."],
    ["Better lines", "Exact rewrites: 'You said X -> say Y instead.'"],
    ["Score /100", "Overall + per-criteria breakdown."],
    ["Next drill", "The single skill + persona to practise next."],
], tone="b", heights=[24] * 6)

r = subheader(s4, r, "C. Readiness / certification gates (Stage 9)", 3)
r = minihead(s4, r, ["Gate", "Must pass"])
r = minirows(s4, r, [
    ["Company knowledge quiz", "OutSkill basics + the funnel."],
    ["Product knowledge quiz", "Accelerator details, pricing & EMI."],
    ["Sales pitch test", "Deliver the pitch end to end."],
    ["Mock-call score", "Minimum score across the ladder incl. the Director round."],
    ["-> Result", "'Sales Ready' badge unlocks live work."],
], tone="y", heights=[24] * 5)

r = subheader(s4, r, "D. Shadowing & live support (Stage 10)", 3)
r = minihead(s4, r, ["Step", "What happens"])
r = minirows(s4, r, [
    ["1. Listen", "Hear recorded senior calls."],
    ["2. Observe", "Watch real customer handling live."],
    ["3. Co-pilot", "Join calls; the senior leads."],
    ["4. Supported live", "Rep leads; senior on standby."],
    ["5. Independent", "Solo live calls."],
], tone="l", heights=[22] * 5)

r = subheader(s4, r, "E. Continuous improvement loop (Stage 12)", 3)
r = minirows(s4, r, [[
    "Call happens  ->  transcript analysed  ->  AI gives feedback  ->  rep drills ONE weak area in the "
    "Mock-Call Room  ->  repeats the mock call  ->  score improves  ->  manager sees progress.   (ongoing)"
]], tone="l", heights=[40], bold0=False)
s4.merge_cells(start_row=r - 2, start_column=1, end_row=r - 2, end_column=3)
widths(s4, [26, 64, 20])
s4.freeze_panes = "A4"


# ================================================================ SHEET 5: SITEMAP & MVP
s5 = wb.create_sheet("5. Sitemap & MVP")
title_block(s5, "Website Sitemap, First-Login & MVP", "How to navigate it and what to build first", 3)
r = subheader(s5, 4, "A. Website navigation (sitemap)", 3)
r = minirows(s5, r, [[
    "Dashboard  ->  My Learning Path  ->  Company Basics  ->  Sales Training  ->  Mock-Call Room  ->  "
    "Feedback  ->  Certification  ->  Live Leads  ->  Reports"
]], tone="g", heights=[34], bold0=False)
s5.merge_cells(start_row=r - 2, start_column=1, end_row=r - 2, end_column=3)

r = subheader(s5, r, "B. First login must answer 3 things instantly", 3)
r = minihead(s5, r, ["Question", "What the screen says"])
r = minirows(s5, r, [
    ["Where am I?", "\"You are in OutSkill onboarding.\""],
    ["What should I do today?", "\"Complete Company Basics - Day 1.\""],
    ["How do I become Sales Ready?", "\"Finish learning -> practice mock calls -> pass certification.\""],
], tone="a", heights=[24, 24, 24])

r = subheader(s5, r, "C. Key screens to build", 3)
r = minirows(s5, r, [
    ["Login / signup", "Welcome dashboard"],
    ["Learning path", "Company overview"],
    ["Sales training module", "Mock-call simulator"],
    ["Feedback page", "Scorecard"],
    ["Certification page", "Live work dashboard"],
], tone="g", heights=[22] * 5)

r = subheader(s5, r, "D. MVP — build fast (5 stages)", 3)
r = minihead(s5, r, ["#", "MVP Stage"])
r = minirows(s5, r, [
    ["1", "Login"],
    ["2", "Training modules"],
    ["3", "Mock-call practice"],
    ["4", "Feedback + score"],
    ["5", "Readiness unlock"],
], tone="y", heights=[20] * 5)
r = minirows(s5, r, [["Note", "This MVP alone already saves a lot of senior time."]],
             tone="y", heights=[24])

r = subheader(s5, r, "E. Suggested build order (what exists vs to add)", 3)
r = minirows(s5, r, [
    ["Already built", "dept -> home -> practice (personas + scorecard + feedback), realcall, reports, followups."],
    ["Add 1", "Auth + profile + welcome dashboard (the spine that tracks progress)."],
    ["Add 2", "Progress/gating model — locked/unlocked stages persisted per user."],
    ["Add 3", "Foundations + funnel quiz, then Mastermind player."],
    ["Add 4", "Department presentations + Sales knowledge modules."],
    ["Add 5", "Level structure on the mock engine + certification gate."],
    ["Add 6", "Shadowing tracker + continuous-improvement loop on real calls."],
], tone="l", heights=[34, 26, 26, 24, 24, 24, 28])
widths(s5, [26, 64, 20])
s5.freeze_panes = "A4"


out = "/home/user/outskill-garvit-/OutSkill_User_Journey.xlsx"
wb.save(out)
print("Saved:", out, "| sheets:", wb.sheetnames)
