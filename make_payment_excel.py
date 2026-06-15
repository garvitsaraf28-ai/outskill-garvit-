#!/usr/bin/env python3
"""Generate an Excel file documenting Outskill's payment gateways / EMI options."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling
TITLE_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FILL  = PatternFill("solid", fgColor="2E75B6")   # blue
INDIA_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green
INTL_FILL    = PatternFill("solid", fgColor="FCE4D6")   # light orange
NOTE_FILL    = PatternFill("solid", fgColor="FFF2CC")   # light yellow

WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, color="FFFFFF", size=16)
SUB_FONT    = Font(italic=True, color="FFFFFF", size=10)
CELL_FONT   = Font(size=10)
BOLD_FONT   = Font(bold=True, size=10)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP_TOP  = Alignment(wrap_text=True, vertical="top",    horizontal="left")
CENTER    = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ---------------------------------------------------------------- workbook
wb = Workbook()
ws = wb.active
ws.title = "Payment Gateways"

headers = [
    "Customer Segment",
    "Payment Mode",
    "Payment Gateway / Platform",
    "Installment Tenure",
    "Extra Charges / Fees",
    "Card / Eligibility Requirement",
    "Process / Notes",
]

rows = [
    # ---- Indian learners -------------------------------------------------
    ["Indian (Indian card / bank)", "One-Time (full amount)", "Razorpay",
     "N/A (full payment)", "No extra charge", "Indian card / bank",
     "Standard one-time full payment.", "india"],

    ["Indian (Indian card / bank)", "Installment (EMI)", "Pine Labs",
     "3 / 6 / 9 / 12 months", "No-cost (zero-cost) EMI — all interest borne by the company",
     "Credit card required", "Share the Pine Labs link with the learner.", "india"],

    ["Indian (no credit card)", "Installment (Debit Card EMI)", "Shopse",
     "3 / 6 / 9 / 12 months", "No-cost (zero-cost) EMI",
     "Debit card; mobile number must pass eligibility check",
     "First check if the learner's number is eligible for debit-card EMI; "
     "only then proceed with Shopse.", "india"],

    ["Indian (no credit card & not eligible for Shopse)", "Installment (via NBFC)",
     "NBFC partners — Fibe / Propel", "As per NBFC", "As per NBFC",
     "Subject to NBFC eligibility",
     "Share the Google form (Fibe or Propel) -> learner fills the form -> a "
     "confirmation page appears -> learner shares the confirmation-page "
     "screenshot -> it is forwarded to the respective team (Fibe / Propel) -> "
     "that team handles the further process. Company's only role: share the "
     "Google form.", "india"],

    # ---- International learners ------------------------------------------
    ["International (non-Indian card)", "One-Time (full amount)", "XP",
     "N/A (full payment)", "No convenience fee", "International card",
     "Standard one-time full payment.", "intl"],

    ["International (non-Indian card)", "Installment", "XP",
     "3 / 6 / 9 / 12 months", "9% convenience fee applies",
     "International card",
     "Installment available up to 12 months; a 9% convenience fee is charged.",
     "intl"],
]

# ---------------------------------------------------------------- title rows
ncols = len(headers)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
t = ws.cell(row=1, column=1, value="Outskill — Payment Gateways & EMI Options")
t.font = TITLE_FONT
t.fill = TITLE_FILL
t.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[1].height = 30

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
s = ws.cell(row=2, column=1,
            value="Payment platforms used for the AI-accelerated program — "
                  "split by Indian vs. International learners")
s.font = SUB_FONT
s.fill = TITLE_FILL
s.alignment = Alignment(vertical="center", horizontal="center")
ws.row_dimensions[2].height = 20

# ---------------------------------------------------------------- header row
HEADER_ROW = 4
for c, head in enumerate(headers, start=1):
    cell = ws.cell(row=HEADER_ROW, column=c, value=head)
    cell.font = WHITE_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[HEADER_ROW].height = 30

# ---------------------------------------------------------------- data rows
for i, row in enumerate(rows):
    r = HEADER_ROW + 1 + i
    region = row[-1]
    fill = INDIA_FILL if region == "india" else INTL_FILL
    for c in range(ncols):
        cell = ws.cell(row=r, column=c + 1, value=row[c])
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = WRAP_TOP
        # bold the first (segment) and third (platform) columns
        cell.font = BOLD_FONT if c in (0, 2) else CELL_FONT
    ws.row_dimensions[r].height = 60

# ---------------------------------------------------------------- column widths
widths = [26, 24, 24, 18, 30, 28, 50]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# freeze panes below header
ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

# ---------------------------------------------------------------- legend / notes
note_row = HEADER_ROW + 1 + len(rows) + 1
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols)
lg = ws.cell(row=note_row, column=1, value="Legend & Key Notes")
lg.font = Font(bold=True, size=11, color="1F3864")

notes = [
    "Green rows = Indian learners (Indian card / bank).   Orange rows = International learners.",
    "Indian one-time full payment -> Razorpay (no extra charge).",
    "Indian installment with credit card -> Pine Labs link; 3/6/9/12 months; zero-cost EMI (company bears interest).",
    "Indian installment without credit card -> Shopse debit-card EMI (zero-cost), only after the number passes the eligibility check.",
    "Indian — no credit card AND not eligible for Shopse -> NBFC route (Fibe / Propel) via Google form + confirmation-page screenshot.",
    "International (one-time or installment) -> XP. Installment up to 12 months carries a 9% convenience fee; one-time has no convenience fee.",
]
for j, text in enumerate(notes, start=1):
    r = note_row + j
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(row=r, column=1, value=("•  " + text))
    cell.font = CELL_FONT
    cell.fill = NOTE_FILL
    cell.alignment = WRAP_TOP
    ws.row_dimensions[r].height = 18

out = "/home/user/outskill-garvit-/Outskill_Payment_Gateways.xlsx"
wb.save(out)
print("Saved:", out)
